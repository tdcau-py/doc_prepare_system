from PyQt5 import QtWidgets, QtCore, uic
import openpyxl
import os


def data_to_xlfile(sheet_name: str, data: dict) -> None:
    """Принимает лист из книги Excel и заполняет его данными из словаря"""
    count = 0
    sheet_cells = sheet_name.merged_cells.ranges  # Выбирает объединенные ячейки

    # формирование списка объединенных ячеек (только сформированные из 3-х ячеек) 
    tmp_sheet = [[*cells.cells] for cells in sheet_cells if len([*cells.cells]) == 3]
    
    rows_sheet = get_rows(sheet_name, tmp_sheet)  # список строк, в которых присутствуют объединенные ячейки
    data_row = sorted([t for t in tmp_sheet if t[0][0] in rows_sheet])  # выбор ячеек, которые относятся к определенной строке
    
    for k, v in data.items():  # перебор словаря с данными и заполнение соответствующих ячеек соответствующей строки
        range_cells = None
        
        if type(k) == tuple:
            for item in k:
                range_cells = [i for i in item]

            for letter in v:
                while True:
                    cell = sheet_name.cell(data_row[count][0][0], data_row[count][0][1])
                    if cell in range_cells:
                        sheet_name[cell.coordinate] = letter.upper()
                        count += 1
                        break
                    else:
                        count += 1

        else:
            for letter in v:
                while True:
                    cell = sheet_name.cell(data_row[count][0][0], data_row[count][0][1])
                    if cell == k:
                        sheet_name[cell.coordinate] = letter.upper()
                        count += 1
                        break
                    else:
                        count += 1


def get_rows(sheet: str, cells: list) -> list:
    """Возвращает список номеров строк листа, в которых присутствуют необходимые для заполнения ячейки"""
    tmp_rows_sheet = set()
    count = 0
    for row in range(len(cells)):
        cell = sheet.cell(cells[count][0][0], cells[count][0][1])
        tmp_rows_sheet.add(cell.row)
        count += 1
    rows = sorted([cell for cell in tmp_rows_sheet])
    return rows


class FormWindow(QtWidgets.QDialog):
    def __init__(self, parent=None):
        super(FormWindow, self).__init__(parent)
        gui, base = uic.loadUiType('forms/notif_form.ui')
        self.ui = gui()
        self.ui.setupUi(self)

        self.setWindowFlags(QtCore.Qt.Window)

        with open('static/style_notification/style.css', 'r') as file:
            style = file.read()

        self.setStyleSheet(style)
        self.setWindowTitle('Форма уведомления о прибытии иностранного гражданина')
        self.showMaximized()

        # Данные мигрантов
        self.mig_lastName = self.ui.lastName_line       # Фамилия
        self.mig_firstName = self.ui.firstName_line     # Имя
        self.mig_patronymic = self.ui.patronymicLine    # Отчество
        self.nationality = self.ui.nationalityLine      # Гражданство
        self.date_birth_mgt = self.ui.dateBirth_mgt     # Дата рождения (мигрант)
        self.mig_male = self.ui.rbtnMale                # Мужской пол
        self.mig_female = self.ui.rbtnFemale            # Женский пол
        self.place_birth_line1 = self.ui.placeBirth_line1
        self.place_birth_line2 = self.ui.placeBirth_line2
        self.place_birth_line3 = self.ui.placeBirth_line3

        # Данные документа, удостоверяющего личность
        self.kind_of_doc = self.ui.kindOfDoc_line       # Вид документа
        self.serial_of_doc = self.ui.serialOfDoc_line
        self.num_of_doc = self.ui.numOfDoc_line
        self.date_issue_doc = self.ui.dateIssue_doc       # Дата выдачи
        self.date_valid_doc = self.ui.dateValid_doc       # Срок действия

        # Данные документа подтверждающего право на пребывание
        self.visa = self.ui.checkVisa
        self.residence_permit = self.ui.checkRsdnc_permit
        self.temporary_permit = self.ui.checkTemp_permit
        self.learning_permit = self.ui.checkLearn_permit
        self.serial_conf_doc = self.ui.serialConfDoc    # Серия подтверждающего документа
        self.num_conf_doc = self.ui.numConfDoc          # Номер подтверждающего документа
        self.date_issue_doc2 = self.ui.dateIssue_doc2
        self.date_valid_doc2 = self.ui.dateValid_doc2

        self.check_service = self.ui.checkService
        self.check_tourism = self.ui.checkTourism
        self.check_business = self.ui.checkBusiness
        self.check_learning = self.ui.checkLearning
        self.check_job = self.ui.checkJob
        self.check_private = self.ui.checkPrivate
        self.check_transit = self.ui.checkTransit
        self.check_humanitarian = self.ui.checkHumanitarian
        self.check_another = self.ui.checkAnother

        self.phone = self.ui.phoneNumber
        self.job = self.ui.jobLine

        # Дата въезда
        self.date_entry = self.ui.dateEntry

        # Срок пребывания до
        self.date_of_stay = self.ui.dateOfStay

        self.mig_card_serial = self.ui.migCard_serial
        self.mig_card_numb = self.ui.migCard_num

        # Сведения о законных представителях
        self.legal_repr1 = self.ui.legalRepr_line1
        self.legal_repr2 = self.ui.legalRepr_line2
        self.legal_repr3 = self.ui.legalRepr_line3

        # Адрес прежнего места пребывания в РФ
        self.past_address_line1 = self.ui.pastAddress_line1
        self.past_address_line2 = self.ui.pastAddress_line2
        self.past_address_line3 = self.ui.pastAddress_line3
        self.past_address_line4 = self.ui.pastAddress_line4

        # Раздел 2. Сведения о месте пребывания
        self.pos_oblast1 = self.ui.posOblast1
        self.pos_oblast2 = self.ui.posOblast2
        self.pos_zone = self.ui.posZone
        self.pos_city = self.ui.posCity
        self.pos_street = self.ui.posStreet
        self.pos_house = self.ui.posHouse
        self.pos_num_house = self.ui.posNumHouse
        self.pos_building_1 = self.ui.posBuilding_1
        self.pos_building_2 = self.ui.posBuilding_2
        self.pos_room = self.ui.posRoom
        self.pos_num_room = self.ui.posNumRoom

        self.dwelling = self.ui.rbtnDwelling
        self.another_room = self.ui.rbtnAnotherRoom
        self.company = self.ui.rbtnCompany

        # Фактическое место нахождения
        self.fact_oblast_line1 = self.ui.factOblast_line1
        self.fact_oblast_line2 = self.ui.factOblast_line2
        self.fact_zone = self.ui.factZone
        self.fact_city = self.ui.factCity
        self.cadastr_num1 = self.ui.cadastrNum_1
        self.cadastr_num2 = self.ui.cadastrNum_2
        self.right_use_line1 = self.ui.rightUse_line1
        self.right_use_line2 = self.ui.rightUse_line2
        self.right_use_line3 = self.ui.rightUse_line3

        # Сведения о принимающей стороне
        self.company_receiv = self.ui.rbtnCompanyReceiv
        self.natural_person = self.ui.rbtnNaturalPerson
        self.last_name_rcvg = self.ui.lastName_rcvg
        self.first_name_rcvg = self.ui.firstName_rcvg
        self.patronymic_rcvg = self.ui.patronymic_rcvg
        self.doc_type_rcvg = self.ui.docType_rcvg
        self.doc_serial_rcvg = self.ui.docSerial_rcvg
        self.doc_number_rcvg = self.ui.docNum_rcvg
        self.date_issue_doc3 = self.ui.dateIssue_doc3
        self.date_valid_doc3 = self.ui.dateValid_doc3
        self.oblast_rcvg_line1 = self.ui.oblastLine1_rcvg
        self.oblast_rcvg_line2 = self.ui.oblastLine2_rcvg
        self.zone_rcvg = self.ui.zoneLine_rcvg
        self.city_rcvg = self.ui.cityLine_rcvg
        self.street_rcvg = self.ui.streetLine_rcvg
        self.house_number = self.ui.houseNum_rcvg
        self.build1_rcvg = self.ui.buildLine1_rcvg
        self.build2_rcvg = self.ui.buildLine2_rcvg
        self.room_rcvg = self.ui.roomLine_rcvg
        self.phone_rcvg = self.ui.phoneNum_rcvg
        self.name_company_line1 = self.ui.nameCompany_1
        self.name_company_line2 = self.ui.nameCompany_2
        self.inn = self.ui.innLine
        self.address_company_line1 = self.ui.companyAddress_1
        self.address_company_line2 = self.ui.companyAddress_2
        self.address_company_line3 = self.ui.companyAddress_3

        btn_generate = self.ui.btnGenerate
        btn_clear = self.ui.btnClear
        self.btn_next = self.ui.btn_next
        self.btn_back = self.ui.btn_back

        self.ui.tabWidget.setCurrentIndex(0)
        self.mig_lastName.setFocus()
        self.btn_back.setDisabled(True)

        btn_generate.clicked.connect(self.on_fill)
        btn_clear.clicked.connect(self.on_clear)
        self.btn_next.clicked.connect(self.on_next)
        self.btn_back.clicked.connect(self.on_back)

        self.ui.tabWidget.currentChanged.connect(self.change_tab)

    def on_fill(self):
        """Принимает данные из формы и заполняет ячейки в документе Excel"""
        date_birth_mgt = self.date_birth_mgt.text().split('.')
        day_birth_mgt = date_birth_mgt[0]
        month_birth_mgt = date_birth_mgt[1]
        year_birth_mgt = date_birth_mgt[2]

        date_issue_doc = self.date_issue_doc.text().split('.')
        day_issue_doc = date_issue_doc[0]
        month_issue_doc = date_issue_doc[1]
        year_issue_doc = date_issue_doc[2]

        date_valid_doc = self.date_valid_doc.text().split('.')
        day_valid_doc = date_valid_doc[0]
        month_valid_doc = date_valid_doc[1]
        year_valid_doc = date_valid_doc[2]

        date_issue_doc2 = self.date_issue_doc2.text().split('.')
        day_issue_doc2 = date_issue_doc2[0]
        month_issue_doc2 = date_issue_doc2[1]
        year_issue_doc2 = date_issue_doc2[2]

        date_valid_doc2 = self.date_valid_doc2.text().split('.')
        day_valid_doc2 = date_valid_doc2[0]
        month_valid_doc2 = date_valid_doc2[1]
        year_valid_doc2 = date_valid_doc2[2]

        date_issue_doc3 = self.date_issue_doc3.text().split('.')
        day_issue_doc3 = date_issue_doc3[0]
        month_issue_doc3 = date_issue_doc3[1]
        year_issue_doc3 = date_issue_doc3[2]

        date_valid_doc3 = self.date_valid_doc3.text().split('.')
        day_valid_doc3 = date_valid_doc3[0]
        month_valid_doc3 = date_valid_doc3[1]
        year_valid_doc3 = date_valid_doc3[2]

        date_entry = self.date_entry.text().split('.')
        day_entry = date_entry[0]
        month_entry = date_entry[1]
        year_entry = date_entry[2]

        date_of_stay = self.date_of_stay.text().split('.')
        day_of_stay = date_of_stay[0]
        month_of_stay = date_of_stay[1]
        year_of_stay = date_of_stay[2]

        gender_male = ' '
        gender_female = ' '

        if self.mig_male.isChecked():
            gender_male = 'X'
        elif self.mig_female.isChecked():
            gender_female = 'X'

        visa = ' '
        residence_permit = ' '
        temp_permit = ' '
        learn_permit = ' '

        if self.visa.isChecked():
            visa = 'X'
        elif self.residence_permit.isChecked():
            residence_permit = 'X'
        elif self.temporary_permit.isChecked():
            temp_permit = 'X'
        elif self.learning_permit.isChecked():
            learn_permit = 'X'

        service = ' '
        tourism = ' '
        business = ' '
        learning = ' '
        job = ' '
        private = ' '
        transit = ' '
        humanitarian = ' '
        another = ' '

        if self.check_service.isChecked():
            service = 'X'
        if self.check_tourism.isChecked():
            tourism = 'X'
        if self.check_business.isChecked():
            business = 'X'
        if self.check_learning.isChecked():
            learning = 'X'
        if self.check_job.isChecked():
            job = 'X'
        if self.check_private.isChecked():
            private = 'X'
        if self.check_transit.isChecked():
            transit = 'X'
        if self.check_humanitarian.isChecked():
            humanitarian = 'X'
        if self.check_another.isChecked():
            another = 'X'

        dwelling = ' '
        another_room = ' '
        company = ' '
        if self.dwelling.isChecked():
            dwelling = 'X'
        elif self.another_room.isChecked():
            another_room = 'X'
        elif self.company.isChecked():
            company = 'X'

        company_rcvg = ' '
        natural_person = ' '
        if self.company_receiv.isChecked():
            company_rcvg = 'X'
        elif self.natural_person.isChecked():
            natural_person = 'X'

        wb = openpyxl.load_workbook('templates/uvedomlenie_templ.xlsx')
        ws = wb.worksheets
        sheet1 = ws[0]  # лист 1
        sheet2 = ws[1]  # лист 2
        sheet3 = ws[2]  # лист 3
        sheet4 = ws[3]  # лист 4

        # Словарь с диапазоном ячеек, которые необходимо заполнить и данными из формы для них
        # Для 1-го листа
        data_client_sh1 = {
            sheet1['N12':'DN12']: self.mig_lastName.text(),
            sheet1['N14':'DN14']: self.mig_firstName.text(),
            sheet1['Z16':'DN16']: self.mig_patronymic.text(),
            sheet1['V18':'DN18']: self.nationality.text(),
            sheet1['AD21':'AH21']: day_birth_mgt, 
            sheet1['AT21':'AX21']: month_birth_mgt, 
            sheet1['BF21':'BR21']: year_birth_mgt, 
            sheet1['CL21']: gender_male, 
            sheet1['DB21']: gender_female,
            sheet1['Z23':'DN23']: self.place_birth_line1.text(),
            sheet1['Z25':'DN25']: self.place_birth_line2.text(),
            sheet1['Z27':'DN27']: self.place_birth_line3.text(),
            sheet1['J29':'AT29']: self.kind_of_doc.text(),
            sheet1['BF29':'BR29']: self.serial_of_doc.text(),
            sheet1['BZ29':'DN29']: self.num_of_doc.text(), 
            sheet1['I31':'M31']: day_issue_doc, 
            sheet1['Z31':'AD31']: month_issue_doc, 
            sheet1['AL31':'AX31']: year_issue_doc, 
            sheet1['BN31':'BR31']: day_valid_doc, 
            sheet1['CD31':'CH31']: month_valid_doc, 
            sheet1['CP31':'DB31']: year_valid_doc,
            sheet1['H36']: visa, 
            sheet1['AJ36']: residence_permit, 
            sheet1['BT36']: temp_permit,
            sheet1['DD36']: learn_permit,
            sheet1['J40':'V40']: self.serial_conf_doc.text(), 
            sheet1['AD40':'CH40']: self.num_conf_doc.text(),
            sheet1['I42':'M42']: day_issue_doc2, 
            sheet1['Z42':'AD42']: month_issue_doc2, 
            sheet1['AL42':'AX42']: year_issue_doc2, 
            sheet1['BN42':'BR42']: day_valid_doc2, 
            sheet1['CD42':'CH42']: month_valid_doc2, 
            sheet1['CP42':'DB42']: year_valid_doc2,
            sheet1['AD44']: service, 
            sheet1['AQ44']: tourism, 
            sheet1['BD44']: business, 
            sheet1['BO44']: learning, 
            sheet1['CA44']: job, 
            sheet1['CN44']: private, 
            sheet1['DB44']: transit,
            sheet1['AD46']: humanitarian, 
            sheet1['AP46']: another, 
            sheet1['CD46':'DN46']: self.phone.text(),
            sheet1['R48':'DN48']: self.job.text(),
            sheet1['I50':'M50']: day_entry, 
            sheet1['Z50':'AD50']: month_entry, 
            sheet1['AL50':'AX50']: year_entry, 
            sheet1['BN50':'BR50']: day_of_stay, 
            sheet1['CD50':'CH50']: month_of_stay, 
            sheet1['CP50':'DB50']: year_of_stay,
            sheet1['AP52':'BB52']: self.mig_card_serial.text(), 
            sheet1['BJ52':'CX52']: self.mig_card_numb.text(),
            sheet1['V54':'DN54']: self.legal_repr1.text(),
            sheet1['V56':'DN56']: self.legal_repr2.text(),
            sheet1['V58':'DN58']: self.legal_repr3.text(),
        }

        # Для 2-го листа
        data_client_sh2 = {
            sheet2['Z3':'DN3']: self.past_address_line1.text(),
            sheet2['Z5':'DN5']: self.past_address_line2.text(),
            sheet2['Z7':'DN7']: self.past_address_line3.text(),
            sheet2['B9':'DN9']: self.past_address_line4.text(),
            sheet2['V14':'DN14']: self.pos_oblast1.text(),
            sheet2['V16':'DN16']: self.pos_oblast2.text(),
            sheet2['V18':'DN18']: self.pos_zone.text(),
            sheet2['Z20':'DN20']: self.pos_city.text(),
            sheet2['V22':'DN22']: self.pos_street.text(),          
            sheet2['AD24':'BF24']: self.pos_num_house.text(),
            sheet2['BR24':'CH24']: self.pos_building_1.text(),
            sheet2['CX24':'DJ24']: self.pos_building_2.text(),
            sheet2['AJ26':'AV26']: self.pos_num_room.text(),
            sheet2['Y28']: dwelling, 
            sheet2['AZ28']: another_room, 
            sheet2['CB28']: company,
            sheet2['AL31':'DN31']: self.right_use_line1.text(),
            sheet2['AL33':'DN33']: self.right_use_line2.text(),
            sheet2['AL35':'DN35']: self.right_use_line3.text(),
            sheet2['Z41':'DN41']: self.fact_oblast_line1.text(),
            sheet2['Z43':'DN43']: self.fact_oblast_line2.text(),
            sheet2['V45':'DN45']: self.fact_zone.text(),
            sheet2['AH47':'DN47']: self.fact_city.text(),
            sheet2['AH50':'DN50']: self.cadastr_num1.text(),
            sheet2['AH52':'DN52']: self.cadastr_num2.text(),
           }

        # Для 3-го листа
        data_client_sh3 = {
            sheet3['CP3']: company_rcvg,
            sheet3['DN3']: natural_person,
            sheet3['N5':'DN5']: self.last_name_rcvg.text(),
            sheet3['N7':'DN7']: self.first_name_rcvg.text(),
            sheet3['AH9':'DN9']: self.patronymic_rcvg.text(),
            sheet3['F11':'AT11']: self.doc_type_rcvg.text(),
            sheet3['BF11':'BR11']: self.doc_serial_rcvg.text(),
            sheet3['BZ11':'DN11']: self.doc_number_rcvg.text(),
            sheet3['I13':'M13']: day_issue_doc3,
            sheet3['Z13':'AD13']: month_issue_doc3,
            sheet3['AL13':'AX13']: year_issue_doc3,
            sheet3['BN13':'BR13']: day_valid_doc3,
            sheet3['CD13':'CH13']: month_valid_doc3,
            sheet3['CP13':'DB13']: year_valid_doc3,
            sheet3['Z17':'DN17']: self.oblast_rcvg_line1.text(),
            sheet3['Z19':'DN19']: self.oblast_rcvg_line2.text(),
            sheet3['V21':'DN21']: self.zone_rcvg.text(),
            sheet3['AD23':'DN23']: self.city_rcvg.text(),
            sheet3['V25':'DN25']: self.street_rcvg.text(),
            sheet3['J27':'V27']: self.house_number.text(),
            sheet3['AH27':'AX27']: self.build1_rcvg.text(),
            sheet3['BN27':'BZ27']: self.build2_rcvg.text(),
            sheet3['CP27':'DB27']: self.room_rcvg.text(),              
            sheet3['N31':'DN31']: self.mig_lastName.text(),
            sheet3['N33':'DN33']: self.mig_firstName.text(),
            sheet3['AH35':'DN35']: self.mig_patronymic.text(),
            sheet3['R37':'DN37']: self.nationality.text(),
            sheet3['AA39':'AE39']: day_birth_mgt, 
            sheet3['AQ39':'AU39']: month_birth_mgt,
            sheet3['BC39':'BO39']: year_birth_mgt,
            sheet3['CL39']: gender_male, 
            sheet3['DB39']: gender_female,            
            sheet3['Z41':'DN41']: self.place_birth_line1.text(),
            sheet3['Z43':'DN43']: self.place_birth_line2.text(),
            sheet3['Z45':'DN45']: self.place_birth_line3.text(),
            sheet3['F47':'AT47']: self.kind_of_doc.text(),
            sheet3['BF47':'BR47']: self.serial_of_doc.text(),
            sheet3['BZ47':'DN47']: self.num_of_doc.text(),
            sheet3['I49':'M49']: day_issue_doc, 
            sheet3['Z49':'AD49']: month_issue_doc, 
            sheet3['AL49':'AX49']: year_issue_doc, 
            sheet3['BN49':'BR49']: day_valid_doc, 
            sheet3['CD49':'CH49']: month_valid_doc, 
            sheet3['CP49':'DB49']: year_valid_doc,
            sheet3['Z53':'DN53']: self.pos_oblast1.text(),
            sheet3['Z55':'DN55']: self.pos_oblast2.text(),
            sheet3['V57':'DN57']: self.pos_zone.text(),
            sheet3['Z59':'DN59']: self.pos_city.text(),
            sheet3['V61':'DN61']: self.pos_street.text(),
            sheet3['AD63':'BF63']: self.pos_num_house.text(),
            sheet3['BR63':'CH63']: self.pos_building_1.text(),
            sheet3['CX63':'DJ63']: self.pos_building_2.text(),
            sheet3['AJ65':'AV65']: self.pos_num_room.text(),
            sheet3['I68':'M68']: day_of_stay, 
            sheet3['AA68':'AE68']: month_of_stay, 
            sheet3['AM68':'AY68']: year_of_stay,
        }

        # Для 4-го листа
        data_client_sh4 = {
            sheet4['Z3':'BJ3']: self.phone_rcvg.text(),
            sheet4['V5':'DN5']: self.name_company_line1.text(),           
            sheet4['B7':'BN7']: self.name_company_line2.text(), 
            sheet4['BZ7':'DN7']: self.inn.text(),
            sheet4['Z9':'DN9']: self.address_company_line1.text(),
            sheet4['B11':'DN11']: self.address_company_line2.text(),
            sheet4['B13':'DN13']: self.address_company_line3.text(),
            sheet4['N27':'DN27']: self.last_name_rcvg.text(),
            sheet4['N29':'DN29']: self.first_name_rcvg.text(),
            sheet4['Z31':'DN31']: self.patronymic_rcvg.text(),
            sheet4['V33':'DN33']: self.name_company_line1.text(),            
            sheet4['B35':'BN35']: self.name_company_line2.text(), 
            sheet4['BZ35':'DN35']: self.inn.text(),    
        }

        # Заполнение ячеек данными, которые вносятся полностью (непосимвольно в отдельную ячейку)
        sheet2['B24'] = self.pos_house.text().upper()
        sheet2['B26'] = self.pos_room.text().upper()
        sheet3['B63'] = self.pos_house.text().upper()
        sheet3['B65'] = self.pos_room.text().upper()

        data_to_xlfile(sheet1, data_client_sh1)  # Заполняет 1-й лист
        data_to_xlfile(sheet2, data_client_sh2)  # Заполняет 2-й лист
        data_to_xlfile(sheet3, data_client_sh3)  # Заполняет 3-й лист
        data_to_xlfile(sheet4, data_client_sh4)  # Заполняет 4-й лист

        home_dir = os.path.expanduser('~')
        folder_name = 'Уведомления о прибытии'
        xl_file = 'Уведомление'

        if self.mig_lastName.text() and self.mig_firstName.text():
            xl_file = f'Уведомление_{self.mig_lastName.text().capitalize()}_{self.mig_firstName.text()[0].upper()}'

        if os.path.exists(home_dir + '\\Desktop'):
            desktop_dir = home_dir + '\\Desktop'
            try:
                wb.save(f'{desktop_dir}\\{folder_name}\\{xl_file}.xlsx')
            except FileNotFoundError:
                os.mkdir(f'{desktop_dir}\\{folder_name}')
                wb.save(f'{desktop_dir}\\{folder_name}\\{xl_file}.xlsx')
            except PermissionError:
                QtWidgets.QMessageBox.information(self, 'Доступ к файлу', 'Закройте файл и повторите попытку.')

        elif os.path.exists(home_dir + '\\Рабочий стол'):
            desktop_dir = home_dir + '\\Рабочий стол'
            try:
                wb.save(f'{desktop_dir}\\{folder_name}\\{xl_file}.xlsx')
            except FileNotFoundError:
                os.mkdir(f'{desktop_dir}\\{folder_name}')
                wb.save(f'{desktop_dir}\\{folder_name}\\{xl_file}.xlsx')
            except PermissionError:
                QtWidgets.QMessageBox.information(self, 'Доступ к файлу', 'Закройте файл и повторите попытку.')
        
        wb.close()

        self.open_file(f'{desktop_dir}\\{folder_name}\\{xl_file}.xlsx')  # Запуск сформированного файла

    def open_file(self, doc_file: str):
        """Открывает сформированный Excel-файл"""
        os.startfile(doc_file)
        

    def on_clear(self):
        """Очищает форму"""
        self.mig_lastName.setText('')
        self.mig_firstName.setText('')
        self.mig_patronymic.setText('')

        self.nationality.setText('')
        self.date_birth_mgt.setText('')
        self.mig_male.setChecked(True)
        self.place_birth_line1.setText('')
        self.place_birth_line2.setText('')
        self.place_birth_line3.setText('')
        self.kind_of_doc.setText('')
        self.serial_of_doc.setText('')
        self.num_of_doc.setText('')
        self.date_issue_doc.setText('')
        self.date_valid_doc.setText('')

        if self.visa.isChecked():
            self.visa.setChecked(False)
        elif self.residence_permit.isChecked():
            self.residence_permit.setChecked(False)
        elif self.temporary_permit.isChecked():
            self.temporary_permit.setChecked(False)
        elif self.learning_permit.isChecked():
            self.learning_permit.setChecked(False)

        self.serial_conf_doc.setText('')
        self.num_conf_doc.setText('')
        self.date_issue_doc2.setText('')
        self.date_valid_doc2.setText('')

        if self.check_service.isChecked():
            self.check_service.setChecked(False)
        if self.check_tourism.isChecked():
            self.check_tourism.setChecked(False)
        if self.check_business.isChecked():
            self.check_business.setChecked(False)
        if self.check_learning.isChecked():
            self.check_learning.setChecked(False)
        if self.check_job.isChecked():
            self.check_job.setChecked(False)
        if self.check_private.isChecked():
            self.check_private.setChecked(False)
        if self.check_transit.isChecked():
            self.check_transit.setChecked(False)
        if self.check_humanitarian.isChecked():
            self.check_humanitarian.setChecked(False)
        if self.check_another.isChecked():
            self.check_another.setChecked(False)

        self.phone.setText('')
        self.job.setText('')
        self.date_entry.setText('')
        self.date_of_stay.setText('')
        self.mig_card_serial.setText('')
        self.mig_card_numb.setText('')
        self.legal_repr1.setText('')
        self.legal_repr2.setText('')
        self.legal_repr3.setText('')

        self.past_address_line1.setText('')
        self.past_address_line2.setText('')
        self.past_address_line3.setText('')
        self.past_address_line4.setText('')
        self.pos_oblast1.setText('')
        self.pos_oblast2.setText('')
        self.pos_zone.setText('')
        self.pos_city.setText('')
        self.pos_street.setText('')
        self.pos_house.setText('')
        self.pos_room.setText('')

        self.pos_num_house.setText('')
        self.pos_building_1.setText('')
        self.pos_building_2.setText('')
        self.pos_num_room.setText('')
        self.fact_oblast_line1.setText('')
        self.fact_oblast_line2.setText('')
        self.fact_zone.setText('')
        self.fact_city.setText('')
        self.cadastr_num1.setText('')
        self.cadastr_num2.setText('')
        self.right_use_line1.setText('')
        self.right_use_line2.setText('')
        self.right_use_line3.setText('')

        self.natural_person.setChecked(True)
        self.last_name_rcvg.setText('')
        self.first_name_rcvg.setText('')
        self.patronymic_rcvg.setText('')
        self.doc_type_rcvg.setText('')
        self.doc_serial_rcvg.setText('')
        self.doc_number_rcvg.setText('')
        self.date_issue_doc3.setText('')
        self.date_valid_doc3.setText('')
        self.oblast_rcvg_line1.setText('')
        self.oblast_rcvg_line2.setText('')
        self.zone_rcvg.setText('')
        self.city_rcvg.setText('')
        self.street_rcvg.setText('')

        self.house_number.setText('')
        self.build1_rcvg.setText('')
        self.build2_rcvg.setText('')
        self.room_rcvg.setText('')
        self.phone_rcvg.setText('')
        self.name_company_line1.setText('')
        self.name_company_line2.setText('')
        self.inn.setText('')
        self.address_company_line1.setText('')
        self.address_company_line2.setText('')
        self.address_company_line3.setText('')
        self.last_name_rcvg.setText('')
        self.first_name_rcvg.setText('')
        self.patronymic_rcvg.setText('')
        self.name_company_line1.setText('')

        self.mig_lastName.setFocus()
        self.ui.tabWidget.setCurrentIndex(0)
        self.btn_back.setDisabled(True)
        self.btn_next.setEnabled(True)

    def on_next(self):
        """Переключает страницы виджета (вперед)"""
        index = self.ui.tabWidget.currentIndex()
        index += 1

        if index > 0:
            self.btn_back.setEnabled(True)

        if index == 2:
            self.btn_next.setDisabled(True)

        self.ui.tabWidget.setCurrentIndex(index)

    def on_back(self):
        """Переключает страницы виджета (назад)"""
        index = self.ui.tabWidget.currentIndex()
        index -= 1

        if index < 2:
            self.btn_next.setEnabled(True)

        if index == 0:
            self.btn_back.setDisabled(True)

        self.ui.tabWidget.setCurrentIndex(index)

    def change_tab(self):
        """Делает активными/неактивными кнопки переключения страниц виджета (при переключении вкладок)"""
        index = self.ui.tabWidget.currentIndex()
        if index == 0:
            self.btn_back.setDisabled(True)
            self.btn_next.setEnabled(True)
        elif index == 1:
            self.btn_back.setEnabled(True)
            self.btn_next.setEnabled(True)
        elif index == 2:
            self.btn_back.setEnabled(True)
            self.btn_next.setDisabled(True)


if __name__ == '__main__':
    import sys

    with open('style.css', 'r') as file:
        style = file.read()

    app = QtWidgets.QApplication(sys.argv)
    app.setStyleSheet(style)
    win = FormWindow()
    win.show()
    sys.exit(app.exec_())
